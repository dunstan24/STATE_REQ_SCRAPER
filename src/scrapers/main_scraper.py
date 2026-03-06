"""
main_scraper.py — Run all state visa requirement scrapers

Urutan eksekusi:
  ACT → NT → NSW → QLD → SA → TAS → VIC → WA

Setiap state:
  1. Memanggil fungsi scrape utama → returns pd.DataFrame
  2. Memanggil export_results(df)  → saves per-state CSV / JSON / XLSX
  3. Jika gagal → skip, lanjut ke state berikutnya

Output:
  Per state : output_scrape/<state>/requirements_<state>.csv / .json / .xlsx
  Combined  : output_scrape/combined/requirements_all_states.csv / .json / .xlsx
  Log file  : main_scraper.log
"""

import importlib
import logging
import os
import sys
import time
import traceback

import pandas as pd

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("main_scraper.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
_COMBINED_DIR = os.path.join(_SCRIPT_DIR, "output_scrape", "combined")


# ── Safe import ───────────────────────────────────────────────────────────────

def _safe_import(module_name):
    try:
        return importlib.import_module(module_name)
    except Exception as e:
        logger.error(f"[IMPORT] Gagal import '{module_name}': {e}")
        return None

act = _safe_import("act_req_scaper")
nt  = _safe_import("nt_req_scraper")
nsw = _safe_import("nsw_req_scraper")
qld = _safe_import("qld_req_scraper")
sa  = _safe_import("sa_req_scraper")
tas = _safe_import("tas_req_scraper")
vic = _safe_import("vic_req_scraper")
wa  = _safe_import("wa_req_scraper")


# ── Per-state runner ──────────────────────────────────────────────────────────

def run_state(state_name, scrape_fn, export_fn):
    """
    Jalankan satu state scraper dengan error handling.
    Return DataFrame jika sukses, None jika gagal.
    """
    logger.info(f"\n{'='*60}")
    logger.info(f"  Starting: {state_name}")
    logger.info(f"{'='*60}")
    start = time.time()

    try:
        df = scrape_fn()

        if df is None or df.empty:
            logger.warning(f"[{state_name}] DataFrame kosong — skip export.")
            return None

        export_fn(df)
        elapsed = time.time() - start
        logger.info(f"[{state_name}] Selesai dalam {elapsed:.1f}s — {len(df)} baris")
        return df

    except Exception:
        elapsed = time.time() - start
        logger.error(f"[{state_name}] GAGAL setelah {elapsed:.1f}s:\n{traceback.format_exc()}")
        return None


# ── ACT ───────────────────────────────────────────────────────────────────────

def scrape_act():
    df_190 = act.scrape_act_subclass(
        subclass=190,
        url_general="https://www.act.gov.au/migration/skilled-migrants/190-nomination-criteria",
        url_overseas="https://www.act.gov.au/migration/skilled-migrants/190-nomination-criteria/overseas-applicant-eligibility",
        url_canberra="https://www.act.gov.au/migration/skilled-migrants/190-nomination-criteria/canberra-resident-applicant-eligibility",
    )
    df_491 = act.scrape_act_subclass(
        subclass=491,
        url_general="https://www.act.gov.au/migration/skilled-migrants/491-nomination-criteria",
        url_overseas="https://www.act.gov.au/migration/skilled-migrants/491-nomination-criteria/491-nomination-overseas-applicant-eligibility",
        url_canberra="https://www.act.gov.au/migration/skilled-migrants/491-nomination-criteria/491-nomination-canberra-resident-applicant-eligibility",
    )
    return pd.concat([df_190, df_491], ignore_index=True)


# ── NSW ───────────────────────────────────────────────────────────────────────

def scrape_nsw():
    df_190 = nsw.scrape_nsw_subclass(
        subclass=190,
        url="https://www.nsw.gov.au/visas-and-migration/skilled-visas/skilled-nominated-visa-subclass-190",
        kw_general="Basic Eligibility",
        kw_details=[
            "Key Steps for Securing NSW Nomination",
            "Understanding Invitation Rounds and the NSW Skills List",
        ],
    )
    df_491 = nsw.scrape_nsw_subclass(
        subclass=491,
        url="https://www.nsw.gov.au/visas-and-migration/skilled-visas/skilled-work-regional-visa-subclass-491",
        kw_general="About NSW Nomination",
        kw_details=[
            "Understanding Invitation Rounds and the NSW Regional Skills List",
            "Key Steps for Securing NSW Nomination",
        ],
    )
    return pd.concat([df_190, df_491], ignore_index=True)


# ── QLD ───────────────────────────────────────────────────────────────────────

def scrape_qld():
    df_onshore  = qld.scrape_qld_pathway("Workers_Living_in_Queensland",
        "https://migration.qld.gov.au/visa-options/skilled-visas/skilled-workers-living-in-queensland")
    df_offshore = qld.scrape_qld_pathway("Workers_Living_Offshore",
        "https://migration.qld.gov.au/visa-options/skilled-visas/skilled-workers-living-offshore")
    df_building = qld.scrape_qld_pathway("Building_and_Construction",
        "https://migration.qld.gov.au/visa-options/skilled-visas/building-and-construction-workers")
    df_uni      = qld.scrape_qld_pathway("University_Graduates",
        "https://www.migration.qld.gov.au/visa-options/skilled-visas/graduates-of-a-queensland-university",
        component_id="component_1540893")
    df_business = qld.scrape_qld_business(
        "https://migration.qld.gov.au/visa-options/skilled-visas/small-business-owners-operating-in-regional-queensland")
    return pd.concat([df_onshore, df_offshore, df_building, df_uni, df_business], ignore_index=True)


# ── SA ────────────────────────────────────────────────────────────────────────

def scrape_sa():
    df_employment = sa.scrape_sa_pathway("Skilled_Employment_in_South_Australia", sa.URL_SA_EMPLOYMENT)
    df_graduates  = sa.scrape_sa_pathway("South_Australian_Graduates",            sa.URL_SA_GRADUATES)
    df_outer      = sa.scrape_sa_pathway("Outer_Regional_Skilled_Employment",     sa.URL_SA_OUTER)
    df_offshore   = sa.scrape_sa_offshore(sa.URL_SA_OFFSHORE)
    return pd.concat([df_employment, df_graduates, df_outer, df_offshore], ignore_index=True)


# ── TAS ───────────────────────────────────────────────────────────────────────

def scrape_tas():
    all_data = tas.scrape_all_pathways(tas.PATHWAY_URLS)
    return tas.build_wide_dataframe(all_data)


# ── Combined export ───────────────────────────────────────────────────────────

def export_combined(results, state_names):
    """
    Export semua state ke satu Excel file dengan sheet terpisah per state.
    Juga export combined CSV dan JSON.

    Sheet name = state code (ACT, NT, NSW, QLD, SA, TAS, VIC, WA)
    """
    os.makedirs(_COMBINED_DIR, exist_ok=True)

    xlsx_path = os.path.join(_COMBINED_DIR, "requirements_all_states.xlsx")
    csv_path  = os.path.join(_COMBINED_DIR, "requirements_all_states.csv")
    json_path = os.path.join(_COMBINED_DIR, "requirements_all_states.json")

    valid_pairs = [
        (name, df)
        for name, df in zip(state_names, results)
        if df is not None and not df.empty
    ]

    if not valid_pairs:
        logger.error("[Combined] Tidak ada data untuk digabungkan.")
        return

    # ── Multi-sheet Excel ─────────────────────────────────────────────────
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            for state_name, df in valid_pairs:
                df.to_excel(writer, sheet_name=state_name, index=False)

        # Apply formatting to each sheet
        wb = load_workbook(xlsx_path)
        for state_name, df in valid_pairs:
            ws = wb[state_name]

            # Header row styling
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            wrap_top    = Alignment(wrap_text=True, vertical="top")

            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # Data rows — wrap text, align top
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap_top

            # Column widths — wide for text-heavy columns, auto for short ones
            for col in ws.columns:
                col_letter  = col[0].column_letter
                header_text = str(col[0].value or "").lower()
                if any(kw in header_text for kw in ["requirement", "general", "detail", "tse", "tsg", "ter", "tbo"]):
                    ws.column_dimensions[col_letter].width = 80
                else:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            # Row heights based on newline count
            for row in ws.iter_rows(min_row=2):
                max_lines = max(
                    (str(cell.value or "").count("\n") + 1 for cell in row),
                    default=1,
                )
                ws.row_dimensions[row[0].row].height = min(max_lines * 15, 409)

        wb.save(xlsx_path)
        logger.info(f"[Combined] XLSX ({len(valid_pairs)} sheets) → {xlsx_path}")

    except Exception as e:
        logger.warning(f"[Combined] Gagal export XLSX: {e}")

    # ── Combined CSV & JSON ───────────────────────────────────────────────
    try:
        combined = pd.concat([df for _, df in valid_pairs], ignore_index=True)
        combined.to_csv(csv_path,   index=False, encoding="utf-8-sig")
        combined.to_json(json_path, orient="records", indent=4)
        logger.info(f"[Combined] CSV  → {csv_path}")
        logger.info(f"[Combined] JSON → {json_path}")
        logger.info(f"[Combined] {len(combined)} total baris dari {len(valid_pairs)} state")
    except Exception as e:
        logger.warning(f"[Combined] Gagal export CSV/JSON: {e}")

    print(f"\n{'='*60}")
    print(f"  COMBINED EXCEL: {len(valid_pairs)} sheets ({', '.join(n for n, _ in valid_pairs)})")
    print(f"  -> {xlsx_path}")
    print(f"{'='*60}")


# ── n8n Webhook trigger ───────────────────────────────────────────────────────
# URL is read from GitHub Actions Secret: N8N_WEBHOOK_URL
# This should point to the /receive-results webhook in your n8n Cloud workflow
# ─────────────────────────────────────────────────────────────────────────────

N8N_WEBHOOK_URL = os.environ.get("N8N_WEBHOOK_URL", "")


def trigger_n8n_webhook(xlsx_path, successful, total):
    """
    POST ke n8n webhook setelah scraping selesai.
    XLSX dikirim sebagai base64 di dalam payload JSON
    agar n8n Cloud bisa langsung attach ke email tanpa akses disk.

    Payload yang dikirim:
      subject      : subject email
      message      : body email
      filename     : nama file attachment
      file_base64  : isi XLSX di-encode base64
    """
    import urllib.request
    import urllib.error
    import json
    import base64

    if not N8N_WEBHOOK_URL:
        logger.warning("[Webhook] N8N_WEBHOOK_URL belum diset — skip trigger.")
        return

    if not xlsx_path or not os.path.exists(xlsx_path):
        logger.warning(f"[Webhook] File tidak ditemukan di: {xlsx_path} — skip trigger.")
        return

    # Read and encode XLSX as base64
    with open(xlsx_path, "rb") as f:
        file_base64 = base64.b64encode(f.read()).decode("utf-8")

    payload = {
        "subject":     f"Scraper Results — {successful}/{total} states berhasil",
        "message":     (
            f"Scraping selesai.\n"
            f"States berhasil : {successful}/{total}\n"
            f"File terlampir  : requirements_all_states.xlsx"
        ),
        "filename":    "requirements_all_states.xlsx",
        "file_base64": file_base64,
    }

    try:
        data = json.dumps(payload).encode("utf-8")
        req  = urllib.request.Request(
            N8N_WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            logger.info(f"[Webhook] n8n triggered — status {resp.status}")
            print(f"  [Webhook] Email dikirim via n8n ✓ (status {resp.status})")

    except urllib.error.URLError as e:
        logger.error(f"[Webhook] Gagal trigger n8n: {e}")
        print(f"  [Webhook] GAGAL — pastikan N8N_WEBHOOK_URL sudah diisi dan workflow aktif.")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    total_start  = time.time()
    state_names  = ["ACT", "NT", "NSW", "QLD", "SA", "TAS", "VIC", "WA"]
    results      = []

    logger.info("=" * 60)
    logger.info("  MULAI: All State Visa Requirements Scraper")
    logger.info("=" * 60)

    # ── ACT ──────────────────────────────────────────────────────────────
    if act:
        results.append(run_state("ACT", scrape_act, act.export_results))
    else:
        logger.warning("[ACT] Skipped — module tidak tersedia.")
        results.append(None)

    # ── NT ───────────────────────────────────────────────────────────────
    if nt:
        results.append(run_state("NT", nt.scrape_nt, nt.export_results))
    else:
        logger.warning("[NT] Skipped — module tidak tersedia.")
        results.append(None)

    # ── NSW ──────────────────────────────────────────────────────────────
    if nsw:
        results.append(run_state("NSW", scrape_nsw, nsw.export_results))
    else:
        logger.warning("[NSW] Skipped — module tidak tersedia.")
        results.append(None)

    # ── QLD ──────────────────────────────────────────────────────────────
    if qld:
        results.append(run_state("QLD", scrape_qld, qld.export_results))
    else:
        logger.warning("[QLD] Skipped — module tidak tersedia.")
        results.append(None)

    # ── SA ───────────────────────────────────────────────────────────────
    if sa:
        results.append(run_state("SA", scrape_sa, sa.export_results))
    else:
        logger.warning("[SA] Skipped — module tidak tersedia.")
        results.append(None)

    # ── TAS ──────────────────────────────────────────────────────────────
    if tas:
        results.append(run_state("TAS", scrape_tas, tas.export_results))
    else:
        logger.warning("[TAS] Skipped — module tidak tersedia.")
        results.append(None)

    # ── VIC ──────────────────────────────────────────────────────────────
    if vic:
        results.append(run_state("VIC", vic.scrape_vic, vic.export_results))
    else:
        logger.warning("[VIC] Skipped — module tidak tersedia.")
        results.append(None)

    # ── WA ───────────────────────────────────────────────────────────────
    if wa:
        results.append(run_state("WA", wa.scrape_wa, wa.export_results))
    else:
        logger.warning("[WA] Skipped — module tidak tersedia.")
        results.append(None)

    # ── Combined output ───────────────────────────────────────────────────
    export_combined(results, state_names)

    # ── Summary ───────────────────────────────────────────────────────────
    total_elapsed = time.time() - total_start
    successful    = sum(1 for r in results if r is not None)

    logger.info(f"\n{'='*60}")
    logger.info(f"  SELESAI — {successful}/8 states OK")
    logger.info(f"  Total waktu: {total_elapsed:.1f}s ({total_elapsed/60:.1f} menit)")
    logger.info("  Status per state:")
    for name, result in zip(state_names, results):
        if result is not None:
            logger.info(f"    {name:5s} : OK ({len(result)} baris)")
        else:
            logger.info(f"    {name:5s} : GAGAL")
    logger.info(f"{'='*60}")

    print(f"\nSelesai! {successful}/8 states berhasil dalam {total_elapsed/60:.1f} menit.")
    print("Lihat main_scraper.log untuk detail lengkap.")

    # ── Trigger n8n → send email ──────────────────────────────────────────
    xlsx_path = os.path.join(_COMBINED_DIR, "requirements_all_states.xlsx")
    trigger_n8n_webhook(xlsx_path, successful, total=8)