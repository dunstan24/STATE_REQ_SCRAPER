""" ACT State Visa Requirements Scraper 

Alur kerja Function:

__main__
  ├── scrape_act_subclass(190, url_general, url_overseas, url_canberra)
  │     ├── fetch_and_parse(url)  ×3            → BeautifulSoup container (col-md-8)
  │     ├── get_clean_text(container)  ×3        → teks terformat dengan hierarki
  │     │     └── _format_li(li)                 → format 1 <li>: teks utama + sub-poin (•)
  │     └── extract_service_fee_from_soup(soup)  → list nilai $xxx dari <li> "service fee"
  │
  ├── scrape_act_subclass(491, ...)              → sama seperti di atas untuk subclass 491
  │
  ├── pd.concat([df_190, df_491])                → gabung jadi 1 DataFrame (2 baris)
  │
  └── export_results(final_df)
        ├── clean_unicode(text)                  → bersihkan \\u00a0, smart quotes, em dash
        ├── df.to_csv()                          → simpan CSV
        ├── df.to_json()                         → simpan JSON
        ├── df.to_excel()                        → simpan XLSX
        └── _format_excel(filepath)              → wrap_text, auto column width, auto row height
"""


import logging
import os
import re
import unicodedata
import pandas as pd
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from playwright_helper import get_page_source_playwright

# Path output relatif terhadap lokasi script ini
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_OUTPUT_DIR = os.path.join(_SCRIPT_DIR, "output_scrape", "act")

# Setup Logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def extract_service_fee_from_soup(soup):

    fees = []

    for li in soup.find_all("li"):
        text = li.get_text(" ", strip=True).lower()
        if "service fee" in text:
            found = re.findall(r"\$[\d,]+(?:\.\d{2})?", text)
            if found:
                fees.extend(found)

    return fees


def fetch_and_parse(url, selector="#main.col-md-8"):
    """Fetch HTML dan return BeautifulSoup container (col-md-8)."""
    logger.info(f"Fetching: {url}")
    html = get_page_source_playwright(
        url=url, wait_for_selector=selector, extra_wait_seconds=3, bypass_cf=True
    )
    if not html:
        logger.error("Gagal mendapatkan HTML. Cloudflare mungkin memblokir.")
        return None

    soup = BeautifulSoup(html, "lxml")
    container = soup.find("div", {"id": "main", "class": "col-md-8"})
    if not container:
        container = soup.find("div", class_="col-md-8")
    return container


def get_clean_text(container):
    """Ekstrak teks dari container dengan format hierarki.

    - Top-level <li>: poin utama, dipisah \n\n
    - Nested <li> (child): sub-poin dengan prefix \u2022
    - Teks <strong> di awal <li> dijadikan label (bold-like prefix)
    """
    if not container:
        return ""

    # Cari semua <ul> / <ol> langsung di dalam container (top-level lists)
    top_lists = container.find_all(["ul", "ol"], recursive=False)
    if not top_lists:
        # Kadang list dibungkus div, cari lebih dalam
        top_lists = container.find_all(["ul", "ol"])

    if top_lists:
        result_points = []

        for top_list in top_lists:
            # Ambil hanya <li> anak langsung dari list ini
            for li in top_list.find_all("li", recursive=False):
                point_text = _format_li(li)
                if point_text:
                    result_points.append(point_text)

        if result_points:
            return "\n\n".join(result_points)

    # Fallback: ambil semua <li> tanpa hierarki
    lines = [
        li.get_text(" ", strip=True)
        for li in container.find_all("li")
        if li.get_text(strip=True)
    ]
    if lines:
        return "\n\n".join(lines)

    # Fallback: jika tidak ada <li> (seperti halaman general)
    all_text = []
    for el in container.find_all(["h2", "h3", "h4", "p", "span", "div", "a"]):
        text = el.get_text(" ", strip=True)
        if text and text not in all_text:
            all_text.append(text)
    if all_text:
        return "\n\n".join(all_text)

    return container.get_text(separator="\n", strip=True)


def _format_li(li):
    """Format satu <li> element: ambil teks langsung + nested list jadi sub-poin."""
    # Ambil teks langsung dari <li> ini (tanpa teks dari nested <ul>/<ol>)
    direct_text_parts = []
    for child in li.children:
        if isinstance(child, Tag):
            if child.name in ["ul", "ol"]:
                continue  # skip nested list, diproses terpisah
            direct_text_parts.append(child.get_text(" ", strip=True))
        else:
            text = child.strip()
            if text:
                direct_text_parts.append(text)

    main_text = " ".join(direct_text_parts).strip()

    # Cari nested <ul>/<ol> di dalam <li> ini
    nested_lists = li.find_all(["ul", "ol"], recursive=False)
    sub_points = []
    for nested in nested_lists:
        for sub_li in nested.find_all("li", recursive=False):
            sub_text = sub_li.get_text(" ", strip=True)
            if sub_text:
                sub_points.append(f"\u2022  {sub_text}")

    if main_text and sub_points:
        return main_text + "\n" + "\n".join(sub_points)
    elif main_text:
        return main_text
    elif sub_points:
        return "\n".join(sub_points)
    return ""


def scrape_act_subclass(subclass, url_general, url_overseas, url_canberra):
    """Scrape ACT nomination criteria untuk subclass tertentu."""
    logger.info(f"=== Scraping ACT Subclass {subclass} ===")

    soup_general = fetch_and_parse(url_general)
    soup_overseas = fetch_and_parse(url_overseas)
    soup_canberra = fetch_and_parse(url_canberra)

    text_general = get_clean_text(soup_general)
    text_overseas = get_clean_text(soup_overseas)
    text_canberra = get_clean_text(soup_canberra)

    # Ekstrak Service Fee dari setiap halaman dengan label nama halaman
    pages = [
        ("overseas", soup_overseas),
        ("canberra", soup_canberra),
        ("general", soup_general),
    ]

    service_fees = []
    for name, soup in pages:
        if soup:
            fees = extract_service_fee_from_soup(soup)
            if fees:
                for fee in fees:
                    service_fees.append(f"{fee} ({name})")

    if service_fees:
        service_fee_value = ", ".join(service_fees)
    else:
        service_fee_value = "-"

    data = {
        "state code": "ACT",
        "state stream": str(subclass),
        "General Requirements": text_general,
        "Requirements Canberra Residents": text_canberra,
        "Requirements overseas": text_overseas,
        "service fee": service_fee_value,
    }

    return pd.DataFrame([data])


def clean_unicode(text):
    """Bersihkan karakter unicode yang tidak diinginkan dari teks."""
    if not isinstance(text, str):
        return text
    # Normalize NFKD: \u00a0 (non-breaking space) → spasi biasa
    text = unicodedata.normalize("NFKD", text)
    # Smart quotes → straight quotes
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    # En dash / em dash → tanda minus biasa
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    # # Hapus newline
    # text = text.replace("\n", " ")
    # Hapus spasi berlebih
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def export_results(df):
    if df is not None and not df.empty:
        # Bersihkan semua kolom string dari karakter unicode
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].apply(clean_unicode)

        os.makedirs(_OUTPUT_DIR, exist_ok=True)

        csv_path = os.path.join(_OUTPUT_DIR, "requirements_act.csv")
        json_path = os.path.join(_OUTPUT_DIR, "requirements_act.json")
        xlsx_path = os.path.join(_OUTPUT_DIR, "requirements_act.xlsx")

        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        df.to_json(json_path, orient="records", indent=4)

        # Export Excel dengan formatting rapi
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
        _format_excel(xlsx_path)

        logger.info(f"Scraping selesai. Data disimpan ke {_OUTPUT_DIR}")
        print("\n--- Preview DataFrame ---")
        print(df[["state code", "state stream", "service fee"]])
    else:
        logger.error("Dataframe kosong.")


def _format_excel(filepath):
    """Format file Excel agar teks multi-line tampil rapi."""
    wb = load_workbook(filepath)
    ws = wb.active

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        header_text = str(col[0].value or "")

        # Kolom teks panjang (requirements) → lebar 80
        # Kolom pendek (state code, stream, fee) → sesuai header
        if any(keyword in header_text.lower() for keyword in ["requirement", "general"]):
            ws.column_dimensions[col_letter].width = 80
        else:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Auto-adjust row heights berdasarkan jumlah \n
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        max_lines = 1
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                line_count = cell.value.count("\n") + 1
                max_lines = max(max_lines, line_count)
        ws.row_dimensions[row[0].row].height = min(max_lines * 15, 409)  # max Excel row height

    wb.save(filepath)
    logger.info(f"[Excel] Formatting selesai: {filepath}")


if __name__ == "__main__":
    # Baris 1: Subclass 190
    df_190 = scrape_act_subclass(
        subclass=190,
        url_general="https://www.act.gov.au/migration/skilled-migrants/190-nomination-criteria",
        url_overseas="https://www.act.gov.au/migration/skilled-migrants/190-nomination-criteria/overseas-applicant-eligibility",
        url_canberra="https://www.act.gov.au/migration/skilled-migrants/190-nomination-criteria/canberra-resident-applicant-eligibility",
    )

    # Baris 2: Subclass 491
    df_491 = scrape_act_subclass(
        subclass=491,
        url_general="https://www.act.gov.au/migration/skilled-migrants/491-nomination-criteria",
        url_overseas="https://www.act.gov.au/migration/skilled-migrants/491-nomination-criteria/491-nomination-overseas-applicant-eligibility",
        url_canberra="https://www.act.gov.au/migration/skilled-migrants/491-nomination-criteria/491-nomination-canberra-resident-applicant-eligibility",
    )

    # Gabungkan jadi satu DataFrame
    final_df = pd.concat([df_190, df_491], ignore_index=True)

    export_results(final_df)
