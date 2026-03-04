"""
Shared utility functions for all state visa requirement scrapers.

Modul ini berisi fungsi-fungsi umum yang dipakai bersama oleh semua
scraper (ACT, NSW, NT, dll.) agar tidak terjadi duplikasi kode.

Fungsi yang tersedia:
  ── Text Formatting ──
  • get_clean_text(container)    — Ekstrak teks dari container BS4 dengan format hierarki
  • format_li(li)                — Format satu <li> element (teks + sub-poin)

  ── Unicode Cleaning ──
  • clean_unicode(text)          — Bersihkan karakter unicode (smart quotes, em dash, dll.)

  ── Data Extraction ──
  • extract_service_fee(soup)    — Cari semua nilai service fee ($xxx)

  ── Export Helpers ──
  • format_excel(filepath)       — Format file Excel (wrap_text, auto-width, auto-height)
  • export_dataframe(df, ...)    — Export DataFrame ke CSV, JSON, dan formatted XLSX
"""

import logging
import os
import re
import unicodedata

from bs4 import Tag
from openpyxl import load_workbook
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)


# ── Text Formatting ──────────────────────────────────────────────────────────


def get_clean_text(container):
    """Ekstrak teks dari container dengan format hierarki.

    - Top-level <li>: poin utama, dipisah \\n\\n
    - Nested <li> (child): sub-poin dengan prefix •
    - Teks <strong> di awal <li> dijadikan label (bold-like prefix)
    """
    if not container:
        return ""

    # Cari semua <ul> / <ol> langsung di dalam container (top-level lists)
    top_lists = container.find_all(["ul", "ol"], recursive=False)
    if not top_lists:
        # Kadang list dibungkus div, cari lebih dalam
        top_lists = container.find_all(["ul", "ol"])

    if top_lists:
        result_points = []

        for top_list in top_lists:
            # Ambil hanya <li> anak langsung dari list ini
            for li in top_list.find_all("li", recursive=False):
                point_text = format_li(li)
                if point_text:
                    result_points.append(point_text)

        if result_points:
            return "\n\n".join(result_points)

    # Fallback: ambil semua <li> tanpa hierarki
    lines = [
        li.get_text(" ", strip=True)
        for li in container.find_all("li")
        if li.get_text(strip=True)
    ]
    if lines:
        return "\n\n".join(lines)

    # Fallback: jika tidak ada <li> (seperti halaman general)
    all_text = []
    for el in container.find_all(["h2", "h3", "h4", "p", "span", "div", "a"]):
        text = el.get_text(" ", strip=True)
        if text and text not in all_text:
            all_text.append(text)
    if all_text:
        return "\n\n".join(all_text)

    return container.get_text(separator="\n", strip=True)


def format_li(li):
    """Format satu <li> element: ambil teks langsung + nested list jadi sub-poin."""
    # Ambil teks langsung dari <li> ini (tanpa teks dari nested <ul>/<ol>)
    direct_text_parts = []
    for child in li.children:
        if isinstance(child, Tag):
            if child.name in ["ul", "ol"]:
                continue  # skip nested list, diproses terpisah
            direct_text_parts.append(child.get_text(" ", strip=True))
        else:
            text = child.strip()
            if text:
                direct_text_parts.append(text)

    main_text = " ".join(direct_text_parts).strip()

    # Cari nested <ul>/<ol> di dalam <li> ini
    nested_lists = li.find_all(["ul", "ol"], recursive=False)
    sub_points = []
    for nested in nested_lists:
        for sub_li in nested.find_all("li", recursive=False):
            sub_text = sub_li.get_text(" ", strip=True)
            if sub_text:
                sub_points.append(f"\u2022  {sub_text}")

    if main_text and sub_points:
        return main_text + "\n" + "\n".join(sub_points)
    elif main_text:
        return main_text
    elif sub_points:
        return "\n".join(sub_points)
    return ""


# ── Unicode Cleaning ─────────────────────────────────────────────────────────


def clean_unicode(text, remove_newlines=False):
    """Bersihkan karakter unicode yang tidak diinginkan dari teks.

    Parameters
    ----------
    text            : str — teks yang akan dibersihkan
    remove_newlines : bool — jika True, ganti \\n dengan spasi (default: False)
    """
    if not isinstance(text, str):
        return text
    # Normalize NFKD: \u00a0 (non-breaking space) → spasi biasa
    text = unicodedata.normalize("NFKD", text)
    # Smart quotes → straight quotes
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    # En dash / em dash → tanda minus biasa
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    # Opsional: hapus newline
    if remove_newlines:
        text = text.replace("\n", " ")
    # Hapus spasi berlebih
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


# ── Data Extraction ──────────────────────────────────────────────────────────


def extract_service_fee(soup, keywords=None):
    """Cari semua nilai service fee ($xxx) dari <li> yang mengandung keyword tertentu.

    Parameters
    ----------
    soup     : BeautifulSoup element
    keywords : list[str] — keyword yang dicari di dalam <li> (default: ["service fee"])
    """
    if keywords is None:
        keywords = ["service fee"]

    fees = []
    if not soup:
        return fees

    for li in soup.find_all("li"):
        text = li.get_text(" ", strip=True).lower()
        if any(kw.lower() in text for kw in keywords):
            found = re.findall(r"\$[\d,]+(?:\.\d{2})?", text)
            if found:
                fees.extend(found)
    return fees


# ── Export Helpers ────────────────────────────────────────────────────────────


def format_excel(filepath, wide_keywords=None):
    """Format file Excel agar teks multi-line tampil rapi.

    Parameters
    ----------
    filepath      : str — path ke file .xlsx
    wide_keywords : list[str] — keyword header yang kolom-nya dibuat lebar 80
                    (default: ["requirement", "general"])
    """
    if wide_keywords is None:
        wide_keywords = ["requirement", "general"]

    wb = load_workbook(filepath)
    ws = wb.active

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        header_text = str(col[0].value or "")

        # Kolom teks panjang (requirements) → lebar 80
        # Kolom pendek (state code, stream, fee) → sesuai header
        if any(keyword in header_text.lower() for keyword in wide_keywords):
            ws.column_dimensions[col_letter].width = 80
        else:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Auto-adjust row heights berdasarkan jumlah \n
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        max_lines = 1
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                line_count = cell.value.count("\n") + 1
                max_lines = max(max_lines, line_count)
        ws.row_dimensions[row[0].row].height = min(
            max_lines * 15, 409
        )  # max Excel row height

    wb.save(filepath)
    logger.info(f"[Excel] Formatting selesai: {filepath}")


def export_dataframe(df, output_dir, filename_prefix, preview_columns=None):
    """Export DataFrame ke CSV, JSON, dan formatted XLSX.

    Parameters
    ----------
    df               : pd.DataFrame
    output_dir       : str — direktori output
    filename_prefix  : str — prefix nama file (misal: "requirements_act")
    preview_columns  : list[str] | None — kolom untuk preview print
                       (default: ["state code", "state stream"])
    """
    if df is None or df.empty:
        logger.error("Dataframe kosong.")
        return

    if preview_columns is None:
        preview_columns = ["state code", "state stream"]

    # Bersihkan semua kolom string dari karakter unicode
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].apply(clean_unicode)

    os.makedirs(output_dir, exist_ok=True)

    csv_path = os.path.join(output_dir, f"{filename_prefix}.csv")
    json_path = os.path.join(output_dir, f"{filename_prefix}.json")
    xlsx_path = os.path.join(output_dir, f"{filename_prefix}.xlsx")

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_json(json_path, orient="records", indent=4)

    # Export Excel dengan formatting rapi
    df.to_excel(xlsx_path, index=False, engine="openpyxl")
    format_excel(xlsx_path)

    logger.info(f"Scraping selesai. Data disimpan ke {output_dir}")
    print("\n--- Preview DataFrame ---")

    # Filter hanya kolom yang ada di DataFrame
    valid_cols = [c for c in preview_columns if c in df.columns]
    if valid_cols:
        print(df[valid_cols])
    else:
        print(df.head())
