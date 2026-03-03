""" NSW State Visa Requirements Scraper

Alur kerja Function:

__main__
  ├── scrape_nsw_subclass(491, url_491)
  │     ├── fetch_and_parse(url)                  → BeautifulSoup container (nsw-layout__main)
  │     ├── get_clean_text(container)              → teks terformat dengan hierarki
  │     │     └── _format_li(li)                   → format 1 <li>: teks utama + sub-poin (•)
  │     └── extract_service_fee_from_soup(soup)    → list nilai $xxx dari <li> "service fee"
  │
  ├── scrape_nsw_subclass(190, url_190)            → sama seperti di atas untuk subclass 190
  │
  ├── pd.concat([df_491, df_190])                  → gabung jadi 1 DataFrame (2 baris)
  │
  └── export_results(final_df)
        ├── clean_unicode(text)                    → bersihkan \\u00a0, smart quotes, em dash
        ├── df.to_csv()                            → simpan CSV
        ├── df.to_json()                           → simpan JSON
        ├── df.to_excel()                          → simpan XLSX
        └── _format_excel(filepath)                → wrap_text, auto column width, auto row height
"""


import logging
import os
import re
import unicodedata
import pandas as pd
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from playwright_helper import get_page_source_playwright

# Path output relatif terhadap lokasi script ini
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_OUTPUT_DIR = os.path.join(_SCRIPT_DIR, "output_scrape", "nsw")

# Setup Logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def extract_service_fee_from_soup(soup):
    """Cari semua nilai service fee ($xxx) dari elemen <li> yang mengandung 'service fee'."""
    fees = []

    for li in soup.find_all("li"):
        text = li.get_text(" ", strip=True).lower()
        if "service fee" in text:
            found = re.findall(r"\$[\d,]+(?:\.\d{2})?", text)
            if found:
                fees.extend(found)

    return fees


def fetch_and_parse(url, selector=".nsw-layout__main"):
    """Fetch HTML dan return BeautifulSoup container (nsw-layout__main)."""
    logger.info(f"Fetching: {url}")
    html = get_page_source_playwright(
        url=url, wait_for_selector=selector, extra_wait_seconds=3, bypass_cf=False
    )
    if not html:
        logger.error("Gagal mendapatkan HTML.")
        return None

    soup = BeautifulSoup(html, "lxml")
    container = soup.find("div", class_="nsw-layout__main")
    return container


    return pd.DataFrame([data])
def extract_general_requirements(container, keyword):
    """Cari div.nsw-wysiwyg-content yang mengandung keyword tertentu,
    lalu ekstrak teks bersih dari div tersebut.

    Parameters
    ----------
    container : BeautifulSoup element (nsw-layout__main)
    keyword   : str — teks pencarian, misal 'Basic Eligibility' atau
                'About NSW Nomination'

    Returns
    -------
    str — teks bersih dari nsw-wysiwyg-content yang cocok, atau '' jika
          tidak ditemukan.
    """
    if not container:
        return ""

    wysiwyg_divs = container.find_all("div", class_="nsw-wysiwyg-content")
    logger.info(f"Ditemukan {len(wysiwyg_divs)} div.nsw-wysiwyg-content")

    for div in wysiwyg_divs:
        text_preview = div.get_text(" ", strip=True)[:200]
        if keyword.lower() in text_preview.lower():
            logger.info(
                f"Matched nsw-wysiwyg-content dengan keyword '{keyword}': "
                f"{text_preview[:80]}..."
            )
            return get_clean_text(div)

    # Fallback: cari heading (h2/h3/h4) yang mengandung keyword,
    # lalu ambil nsw-wysiwyg-content terdekat sesudahnya
    for heading in container.find_all(["h2", "h3", "h4"]):
        if keyword.lower() in heading.get_text(strip=True).lower():
            # Cari nsw-wysiwyg-content sesudah heading ini
            sibling = heading.find_next("div", class_="nsw-wysiwyg-content")
            if sibling:
                logger.info(
                    f"Matched via heading '{heading.get_text(strip=True)}' → "
                    f"nsw-wysiwyg-content berikutnya"
                )
                return get_clean_text(sibling)

    logger.warning(f"Tidak ditemukan nsw-wysiwyg-content dengan keyword '{keyword}'")
    return ""


def extract_li_from_wysiwyg(container, li_keyword):
    """Cari <li> di dalam div.nsw-wysiwyg-content yang mengandung keyword,
    lalu return teks dari <li> tersebut beserta sub-poin-nya.

    Parameters
    ----------
    container  : BeautifulSoup element (nsw-layout__main)
    li_keyword : str — teks yang dicari di dalam <li>, misal 'Residency'

    Returns
    -------
    str — teks bersih dari <li> yang cocok (termasuk nested sub-poin),
          atau '' jika tidak ditemukan.
    """
    if not container:
        return ""

    wysiwyg_divs = container.find_all("div", class_="nsw-wysiwyg-content")

    for div in wysiwyg_divs:
        for li in div.find_all("li"):
            li_text = li.get_text(" ", strip=True)
            if li_keyword.lower() in li_text.lower():
                logger.info(
                    f"Matched <li> dengan keyword '{li_keyword}': "
                    f"{li_text[:80]}..."
                )
                return _format_li(li)

    logger.warning(f"Tidak ditemukan <li> dengan keyword '{li_keyword}'")
    return ""


def get_clean_text(container):
    """Ekstrak teks dari container dengan format hierarki.

    - Top-level <li>: poin utama, dipisah \\n\\n
    - Nested <li> (child): sub-poin dengan prefix •
    - Teks <strong> di awal <li> dijadikan label (bold-like prefix)
    """
    if not container:
        return ""

    # Cari semua <ul> / <ol> langsung di dalam container (top-level lists)
    top_lists = container.find_all(["ul", "ol"], recursive=False)
    if not top_lists:
        # Kadang list dibungkus div, cari lebih dalam
        top_lists = container.find_all(["ul", "ol"])

    if top_lists:
        result_points = []

        for top_list in top_lists:
            # Ambil hanya <li> anak langsung dari list ini
            for li in top_list.find_all("li", recursive=False):
                point_text = _format_li(li)
                if point_text:
                    result_points.append(point_text)

        if result_points:
            return "\n\n".join(result_points)

    # Fallback: ambil semua <li> tanpa hierarki
    lines = [
        li.get_text(" ", strip=True)
        for li in container.find_all("li")
        if li.get_text(strip=True)
    ]
    if lines:
        return "\n\n".join(lines)

    # Fallback: jika tidak ada <li> (seperti halaman general)
    all_text = []
    for el in container.find_all(["h2", "h3", "h4", "p", "span", "div", "a"]):
        text = el.get_text(" ", strip=True)
        if text and text not in all_text:
            all_text.append(text)
    if all_text:
        return "\n\n".join(all_text)

    return container.get_text(separator="\n", strip=True)


def _format_li(li):
    """Format satu <li> element: ambil teks langsung + nested list jadi sub-poin."""
    # Ambil teks langsung dari <li> ini (tanpa teks dari nested <ul>/<ol>)
    direct_text_parts = []
    for child in li.children:
        if isinstance(child, Tag):
            if child.name in ["ul", "ol"]:
                continue  # skip nested list, diproses terpisah
            direct_text_parts.append(child.get_text(" ", strip=True))
        else:
            text = child.strip()
            if text:
                direct_text_parts.append(text)

    main_text = " ".join(direct_text_parts).strip()

    # Cari nested <ul>/<ol> di dalam <li> ini
    nested_lists = li.find_all(["ul", "ol"], recursive=False)
    sub_points = []
    for nested in nested_lists:
        for sub_li in nested.find_all("li", recursive=False):
            sub_text = sub_li.get_text(" ", strip=True)
            if sub_text:
                sub_points.append(f"\u2022  {sub_text}")

    if main_text and sub_points:
        return main_text + "\n" + "\n".join(sub_points)
    elif main_text:
        return main_text
    elif sub_points:
        return "\n".join(sub_points)
    return ""


def scrape_nsw_subclass(subclass, url, kw_general, kw_state, kw_overseas,
                        li_kw_state=None, li_kw_overseas=None):
    """Scrape NSW visa requirements untuk subclass tertentu.

    Parameters
    ----------
    subclass      : int — 190 atau 491
    url           : str — URL halaman NSW
    kw_general    : str — keyword nsw-wysiwyg-content untuk General Requirements
    kw_state      : str — keyword nsw-wysiwyg-content untuk Requirements state Residents
    kw_overseas   : str — keyword nsw-wysiwyg-content untuk Requirements overseas
    li_kw_state   : str | None — jika diisi, ambil Requirements state Residents dari
                    <li> spesifik yang mengandung keyword ini.
    li_kw_overseas: str | None — jika diisi, ambil Requirements overseas dari
                    <li> spesifik yang mengandung keyword ini.
    """
    logger.info(f"=== Scraping NSW Subclass {subclass} ===")

    soup = fetch_and_parse(url)

    # Ekstrak teks dari nsw-wysiwyg-content berdasarkan keyword masing-masing
    text_general = extract_general_requirements(soup, kw_general)

    # State Residents: dari seluruh wysiwyg-content ATAU dari <li> spesifik
    if li_kw_state:
        text_state = extract_li_from_wysiwyg(soup, li_kw_state)
    else:
        text_state = extract_general_requirements(soup, kw_state)

    # Overseas: dari seluruh wysiwyg-content ATAU dari <li> spesifik
    if li_kw_overseas:
        text_overseas = extract_li_from_wysiwyg(soup, li_kw_overseas)
    else:
        text_overseas = extract_general_requirements(soup, kw_overseas)

    # Ekstrak Service Fee
    service_fees = []
    if soup:
        fees = extract_service_fee_from_soup(soup)
        if fees:
            service_fees.extend(fees)

    service_fee_value = ", ".join(service_fees) if service_fees else "-"

    data = {
        "state code": "NSW",
        "state stream": str(subclass),
        "General Requirements": text_general,
        "Requirements state Residents": text_state,
        "Requirements overseas": text_overseas,
        "service fee": service_fee_value,
    }

    return pd.DataFrame([data])


def clean_unicode(text):
    """Bersihkan karakter unicode yang tidak diinginkan dari teks."""
    if not isinstance(text, str):
        return text
    # Normalize NFKD: \u00a0 (non-breaking space) → spasi biasa
    text = unicodedata.normalize("NFKD", text)
    # Smart quotes → straight quotes
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    # En dash / em dash → tanda minus biasa
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    # Hapus spasi berlebih
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def export_results(df):
    if df is not None and not df.empty:
        # Bersihkan semua kolom string dari karakter unicode
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].apply(clean_unicode)

        os.makedirs(_OUTPUT_DIR, exist_ok=True)

        csv_path = os.path.join(_OUTPUT_DIR, "requirements_nsw.csv")
        json_path = os.path.join(_OUTPUT_DIR, "requirements_nsw.json")
        xlsx_path = os.path.join(_OUTPUT_DIR, "requirements_nsw.xlsx")

        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        df.to_json(json_path, orient="records", indent=4)

        # Export Excel dengan formatting rapi
        df.to_excel(xlsx_path, index=False, engine="openpyxl")
        _format_excel(xlsx_path)

        logger.info(f"Scraping selesai. Data disimpan ke {_OUTPUT_DIR}")
        print("\n--- Preview DataFrame ---")
        print(df[["state code", "state stream", "service fee"]])
    else:
        logger.error("Dataframe kosong.")


def _format_excel(filepath):
    """Format file Excel agar teks multi-line tampil rapi."""
    wb = load_workbook(filepath)
    ws = wb.active

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        header_text = str(col[0].value or "")

        # Kolom teks panjang (requirements) → lebar 80
        # Kolom pendek (state code, stream, fee) → sesuai header
        if any(keyword in header_text.lower() for keyword in ["requirement", "general"]):
            ws.column_dimensions[col_letter].width = 80
        else:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Auto-adjust row heights berdasarkan jumlah \n
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        max_lines = 1
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                line_count = cell.value.count("\n") + 1
                max_lines = max(max_lines, line_count)
        ws.row_dimensions[row[0].row].height = min(max_lines * 15, 409)  # max Excel row height

    wb.save(filepath)
    logger.info(f"[Excel] Formatting selesai: {filepath}")


if __name__ == "__main__":
    # Baris 1: Subclass 491
    df_491 = scrape_nsw_subclass(
        subclass=491,
        url="https://www.nsw.gov.au/visas-and-migration/skilled-visas/skilled-work-regional-visa-subclass-491",
        kw_general="About NSW Nomination",
        kw_state="Eligibility Criteria",
        kw_overseas="Key Steps for Securing NSW Nomination",
    )

    # Baris 2: Subclass 190 — state/overseas dari <li> spesifik
    df_190 = scrape_nsw_subclass(
        subclass=190,
        url="https://www.nsw.gov.au/visas-and-migration/skilled-visas/skilled-nominated-visa-subclass-190",
        kw_general="Basic Eligibility",
        kw_state=None,
        kw_overseas=None,
        li_kw_state="Residing in NSW and have continuously resided in NSW",
        li_kw_overseas="offshore",
    )

    # Gabungkan jadi satu DataFrame
    final_df = pd.concat([df_491, df_190], ignore_index=True)

    export_results(final_df)
